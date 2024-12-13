from openpyxl import load_workbook
from tabulate import tabulate

# Función para comparar archivos Excel
def comparar_excels():
    # Solicitar información al usuario
    ruta1 = input("Ingrese el nombre del primer archivo (incluya la extensión, ej: Archivo1.xlsx): ").strip()
    ruta2 = input("Ingrese el nombre del segundo archivo (incluya la extensión, ej: Archivo2.xlsx): ").strip()
    hoja1 = input("Ingrese el nombre de la hoja en el primer archivo: ").strip()
    hoja2 = input("Ingrese el nombre de la hoja en el segundo archivo: ").strip()
    col_comparar_1 = input("Ingrese la columna de los números en el primer archivo (ej: A): ").strip().upper()
    col_comparar_2 = input("Ingrese la columna de los números en el segundo archivo (ej: A): ").strip().upper()
    col_verificar_1 = input("Ingrese la columna de los valores a verificar en el primer archivo (ej: B): ").strip().upper()
    col_verificar_2 = input("Ingrese la columna de los valores a verificar en el segundo archivo (ej: B): ").strip().upper()

    # Cargar los libros y hojas
    libro1 = load_workbook(ruta1, data_only=True)
    libro2 = load_workbook(ruta2, data_only=True)
    hoja_1 = libro1[hoja1]
    hoja_2 = libro2[hoja2]

    # Leer las columnas como listas
    col1_numeros = [hoja_1[f"{col_comparar_1}{i}"].value for i in range(2, hoja_1.max_row + 1)]
    col2_numeros = [hoja_2[f"{col_comparar_2}{i}"].value for i in range(2, hoja_2.max_row + 1)]
    col1_verificar = [hoja_1[f"{col_verificar_1}{i}"].value for i in range(2, hoja_1.max_row + 1)]
    col2_verificar = [hoja_2[f"{col_verificar_2}{i}"].value for i in range(2, hoja_2.max_row + 1)]

    # Comparar y encontrar diferencias
    resultados = []
    for idx, num1 in enumerate(col1_numeros):
        if num1 in col2_numeros:
            idx2 = col2_numeros.index(num1)
            valor1 = col1_verificar[idx]
            valor2 = col2_verificar[idx2]
            if valor1 != valor2:
                resultados.append([
                    num1,
                    valor1,
                    valor2,
                    f"{col_verificar_1}{idx+2} ({ruta1})",
                    f"{col_verificar_2}{idx2+2} ({ruta2})"
                ])

    # Mostrar resultados
    if resultados:
        print("\nDiferencias encontradas:")
        print(tabulate(resultados, headers=["Número", "Valor Archivo1", "Valor Archivo2", "Celda Archivo1", "Celda Archivo2"]))
    else:
        print("\nNo se encontraron diferencias.")

# Ejecutar función
comparar_excels()
